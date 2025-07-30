import os
import cv2
from datetime import datetime
from collections import deque
from openpyxl import Workbook, load_workbook
from PIL import Image, ImageDraw, ImageFont
import qrcode

class BarberShop:
    def __init__(self):
        self.jenis_layanan = {
            1: {"nama_layanan": "Potong Rambut", "harga": 20000},
            2: {"nama_layanan": "Cukur Jenggot", "harga": 15000},
            3: {"nama_layanan": "Cuci Rambut", "harga": 25000},
            4: {"nama_layanan": "Treatment Rambut", "harga": 50000},
            5: {"nama_layanan": "Creambath", "harga": 60000},
            6: {"nama_layanan": "Hair Coloring", "harga": 120000},
            7: {"nama_layanan": "Hair Spa", "harga": 70000},
            8: {"nama_layanan": "Facial", "harga": 40000},
        }
        self.antrian = deque()
        self.kursi = [None, None]

    def get_kursi_kosong(self):
        for i, kursi in enumerate(self.kursi):
            if kursi is None:
                return i
        return None
    
    def is_kursi_penuh(self):
        return all(kursi is not None for kursi in self.kursi)
    
    def kosongkan_kursi(self, transaksi_id):
        for i in range(len(self.kursi)):
            if self.kursi[i] and self.kursi[i] == transaksi_id:
                self.kursi[i] = None
                print(f"Kursi {i+1} telah dikosongkan untuk transaksi {transaksi_id}")
                return True
        return False

    def tampilkan_layanan(self):
        print("\nDaftar Layanan Tersedia:")
        for key, value in self.jenis_layanan.items():
            print(f"{key}. {value['nama_layanan']} - Rp{value['harga']:,}")

    def pesan_layanan(self):
        nama = input("Masukkan nama Anda: ")
        self.tampilkan_layanan()
        
        pilihan_input = input("Masukkan nomor layanan (pisahkan dengan koma/spasi): ")

        pilihan_list = []
        for p in pilihan_input.replace(',', ' ').split():
            if p.isdigit():
                pilihan_list.append(int(p))
        
        layanan_terpilih = []
        for p in pilihan_list:
            if p in self.jenis_layanan:
                layanan = self.jenis_layanan[p]
                layanan_terpilih.append(layanan)
                print(f"✓ {layanan['nama_layanan']} ditambahkan.")
            else:
                print(f"❌ Nomor layanan {p} tidak valid.")
        
        if layanan_terpilih:
            self.antrian.append({
                "nama": nama,
                "layanan": layanan_terpilih
            })
            print(f"\n Pelanggan '{nama}' berhasil ditambahkan ke antrian.")
        else:
            print(" Tidak ada layanan yang dipilih.")

    def tampilkan_antrian(self):
        if not self.antrian:
            print("Antrian kosong.")
            return
        print("\nAntrian Pelanggan:")
        for idx, pelanggan in enumerate(self.antrian, start=1):
            print(f"{idx}. {pelanggan['nama']}")
            for layanan in pelanggan['layanan']:
                print(f"   - {layanan['nama_layanan']} - Rp{layanan['harga']:,}")
        
        print("\nStatus Kursi:")
        for i, kursi in enumerate(self.kursi, start=1):
            status = "Terisi" if kursi else "Kosong"
            print(f"Kursi {i}: {status}")

    def simpan_ke_excel(self, nama, layanan_list, total, status="Belum Selesai"):
        file_name = "transaksi_barbershop.xlsx"
        headers = ["ID Transaksi", "Waktu", "Nama Pelanggan", "Layanan", "Harga", "Status"]

        if os.path.exists(file_name):
            wb = load_workbook(file_name)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(headers)

        last_id = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and str(row[0]).startswith("TRX"):
                try:
                    last_id = max(last_id, int(str(row[0])[3:]))
                except:
                    pass

        transaksi_id = f"TRX{last_id + 1:04d}"
        waktu = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        for layanan in layanan_list:
            ws.append([transaksi_id, waktu, nama, layanan['nama_layanan'], layanan['harga'], status])

        wb.save(file_name)
        print(f"Transaksi {transaksi_id} disimpan ke {file_name}")
        return transaksi_id, waktu

    def buat_struk_gambar(self, transaksi_id, nama, layanan_list, total, waktu):
        qr_data = f"ID: {transaksi_id}\nNama: {nama}\nTotal: Rp{total:,}\nStatus: Selesai"
        qr = qrcode.QRCode(box_size=3, border=2)
        qr.add_data(qr_data)
        qr.make(fit=True)
        qr_img = qr.make_image(fill_color="black", back_color="white")

        width, height = 400, 350 + len(layanan_list) * 30
        img = Image.new("RGB", (width, height), color=(255, 255, 255))
        draw = ImageDraw.Draw(img)

        try:
            font = ImageFont.truetype("arial.ttf", 16)
            font_bold = ImageFont.truetype("arialbd.ttf", 18)
        except:
            font = ImageFont.load_default()
            font_bold = ImageFont.load_default()

        y = 20
        draw.text((20, y), "STRUK RUSDI BARBERSHOP", font=font_bold, fill="black")
        y += 30
        draw.text((20, y), f"ID Transaksi: {transaksi_id}", font=font, fill="black")
        y += 20
        draw.text((20, y), f"Nama: {nama}", font=font, fill="black")
        y += 20
        draw.text((20, y), f"Waktu: {waktu}", font=font, fill="black")
        y += 30
        draw.text((20, y), "Layanan:", font=font_bold, fill="black")
        y += 25

        for layanan in layanan_list:
            draw.text((30, y), f"- {layanan['nama_layanan']}", font=font, fill="black")
            draw.text((250, y), f"Rp{layanan['harga']:,}", font=font, fill="black")
            y += 25

        y += 10
        draw.line([(20, y), (380, y)], fill="black", width=1)
        y += 10
        draw.text((20, y), "Total:", font=font_bold, fill="black")
        draw.text((250, y), f"Rp{total:,}", font=font_bold, fill="black")
        y += 40
        draw.text((80, y), "Terima kasih telah berkunjung!", font=font, fill="black")

        qr_pos = (width - qr_img.size[0] - 20, height - qr_img.size[1] - 20)
        img.paste(qr_img, qr_pos)

        filename = f"struk_{transaksi_id}.png"
        img.save(filename)
        print(f"Struk disimpan dengan QR: {filename}")

    def layani_pelanggan(self):
        if not self.antrian:
            print("Tidak ada pelanggan dalam antrian.")
            return
        
        if self.is_kursi_penuh():
            print("Semua kursi sedang terisi. Tidak bisa melayani pelanggan baru saat ini.")
            return
            
        kursi_index = self.get_kursi_kosong()
        if kursi_index is None:
            print("Tidak ada kursi kosong yang tersedia.")
            return
            
        pelanggan = self.antrian.popleft()
        nama = pelanggan['nama']
        layanan = pelanggan['layanan']
        total = sum(item['harga'] for item in layanan)

        print(f"\n Melayani {nama} di Kursi {kursi_index + 1}...")
        for item in layanan:
            print(f" - {item['nama_layanan']} - Rp{item['harga']:,}")
        print(f"Total: Rp{total:,}")
        print(f"Terima kasih {nama}!\n")

        transaksi_id, waktu = self.simpan_ke_excel(nama, layanan, total, "Dalam Proses")
        self.kursi[kursi_index] = transaksi_id  
        self.buat_struk_gambar(transaksi_id, nama, layanan, total, waktu)

    def extract_transaksi_id(self, qr_text):
        for line in qr_text.split('\n'):
            if line.strip().startswith("ID:"):
                return line.split("ID:")[1].strip()
        return None

    def update_status_excel(self, transaksi_id):
        file_name = "transaksi_barbershop.xlsx"
        try:
            wb = load_workbook(file_name)
            ws = wb.active
            updated = False
            for row in ws.iter_rows(min_row=2):
                if str(row[0].value).strip() == str(transaksi_id).strip():  # Gunakan str() dan strip()
                    row[5].value = "Selesai"
                    updated = True
            if updated:
                wb.save(file_name)
                print(f"Status transaksi {transaksi_id} diperbarui ke Selesai.")
                self.kosongkan_kursi(transaksi_id)
            else:
                print(f"ID transaksi {transaksi_id} tidak ditemukan.")
        except Exception as e:
            print(f"Error: {e}")
            file_name = "transaksi_barbershop.xlsx"
            try:
                wb = load_workbook(file_name)
                ws = wb.active
                updated = False
                for row in ws.iter_rows(min_row=2):
                    if row[0].value == transaksi_id:
                        row[5].value = "Selesai"
                        updated = True
                if updated:
                    wb.save(file_name)
                    print(f"Status transaksi {transaksi_id} Selesai.")
                    self.kosongkan_kursi(transaksi_id)
                else:
                    print(f"ID transaksi {transaksi_id} tidak ditemukan.")
            except FileNotFoundError:
                print("File Excel tidak ditemukan.")

    def scan_qr_pelanggan(self):
        print("Arahkan kamera ke QR Code pelanggan...")
        cap = cv2.VideoCapture(0)
        detector = cv2.QRCodeDetector()

        while True:
            ret, frame = cap.read()
            if not ret:
                continue

            data, bbox, _ = detector.detectAndDecode(frame)
            if data:
                print("QR Terbaca:\n", data)
                cap.release()
                cv2.destroyAllWindows()
                transaksi_id = self.extract_transaksi_id(data)
                if transaksi_id:
                    self.update_status_excel(transaksi_id)
                else:
                    print("ID transaksi tidak ditemukan dalam QR.")
                return

            cv2.imshow("Scan QR Pelanggan", frame)
            if cv2.waitKey(1) & 0xFF == ord('q'):
                break

        cap.release()
        cv2.destroyAllWindows()
        print("Tidak ada QR Code yang terbaca.")

    def mulai(self):
        while True:
            print("\n=== SELAMAT DATANG DI RUSDI BARBERSHOP ===")
            print("\n=== MENU BARBERSHOP ===")
            print("1. Tambah Pelanggan")
            print("2. Tampilkan Antrian")
            print("3. Layani Pelanggan")
            print("4. Scan QR Pelanggan")
            print("5. Keluar")
            pilihan = input("Pilih menu: ")
            if pilihan == '1':
                self.pesan_layanan()
            elif pilihan == '2':
                self.tampilkan_antrian()
            elif pilihan == '3':
                self.layani_pelanggan()
            elif pilihan == '4':
                self.scan_qr_pelanggan()
            elif pilihan == '5':
                print("Terima Kasih!")
                break
            else:
                print("Pilihan tidak valid.")

if __name__ == "__main__":
    app = BarberShop()
    app.mulai()